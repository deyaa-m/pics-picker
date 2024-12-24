import requests  
from bs4 import BeautifulSoup  
import xlsxwriter # type: ignore
import openpyxl
import os


input_dir = os.environ['inputs_dir']
outputs_dir = os.environ['outputs_dir']


# directory = os.fsencode(inputs_dir)
    
for file in os.listdir(input_dir):
    # filename = os.fsdecode(file)
    if file.endswith(".xlsx") : 
        file_name = os.path.splitext(file)[0]
        # print(os.path.join(directory, filename))
    
        source_file = input_dir + file_name + '.xlsx'
        column = 1
        row = 2
        destination_file = outputs_dir + file_name + '.xlsx'
        pics_path = "images/"
        search_link = "https://www.worldofsweets.de/index.php?stoken=8EE4C6B&sid=d7aa5deb7e79f0d40b7b8a8a7afdbc56&lang=0&queryFromSuggest=&userInput=&listorderby=relevance&listorder=desc&cl=fcfatsearch_productlist&searchparam="
        
        wrkbk = openpyxl.load_workbook(source_file) 
        
        sh = wrkbk.active 
        
        def getdata(url):  
            r = requests.get(url)  
            return r.text  
            
        workbook = xlsxwriter.Workbook(destination_file)
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format()
        cell_format.set_align('vcenter')
        
        for j in range(row, sh.max_row+1): 
            product_serial = sh.cell(row=j, column=column) 
            product_name = sh.cell(row=j, column=column + 1) 
            product_url = search_link + str(product_serial.value)
        
            print(product_url)
            
            htmldata = getdata(product_url)  
            soup = BeautifulSoup(htmldata, 'html.parser') 
            images = soup.find_all('img') 
        
            for item in images: 
                image = item['src']
                # try:
                if ".jpg" in image:
                    location = "images/" + image.split("/")[-1]
                    img_data = requests.get(image).content
                    with open(location, 'wb') as handler:
                        handler.write(img_data)
                    print(location)
                    # Insert an image.
                    worksheet.write(j, column, str(product_serial.value), cell_format)
                    worksheet.write(j, column+1, str(product_name.value), cell_format)
                    worksheet.set_column(column+2, 60)
                    worksheet.set_row(j, 120)
                    worksheet.insert_image(j, column+2, location, {"x_scale": 0.25, "y_scale": 0.25})
                    worksheet.write(j, column+3, str(image), cell_format)
                    worksheet.autofit()
            
        workbook.close()
        # except:
            # pass

# driver.quit()
